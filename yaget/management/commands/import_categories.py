# -*- coding: utf-8 -*-
"""
Import Amazon BTG and Wowma category masters into DB.

Usage:
  python manage.py import_categories --source amazon --truncate
  python manage.py import_categories --source wowma --truncate
  python manage.py import_categories --source amazon --path /path/to/dir

Notes:
- Amazon: expects BTG .xls files in category/amazon/. Uses sheet index 1, columns
  "Node ID" and "Node Path". Levels are split by '/'.
- Wowma: expects category.xlsx in category/wowma/. Columns:
  出品カテゴリID, カテゴリフルパス, 第一階層名, 第二階層名, 第三階層名, 第四階層名.
"""
import argparse
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.conf import settings

import xlrd
import openpyxl

from yaget.models import AmaCategory, WowCategory


class Command(BaseCommand):
    help = "Import Amazon/Wowma categories from local Excel files into DB."

    def add_arguments(self, parser):
        parser.add_argument("--source", choices=["amazon", "wowma"], required=True)
        parser.add_argument("--path", type=str, help="Directory containing source files")
        parser.add_argument(
            "--truncate",
            action="store_true",
            help="Delete existing records before import",
        )
        parser.add_argument(
            "--limit",
            type=int,
            help="Limit number of rows (for testing)",
        )

    def handle(self, *args, **options):
        source = options["source"]
        base_dir = Path(options["path"] or (Path(settings.BASE_DIR) / "category" / source))

        if not base_dir.exists():
            raise CommandError(f"Source path not found: {base_dir}")

        if source == "amazon":
            if options["truncate"]:
                AmaCategory.objects.all().delete()
                self.stdout.write("Truncated AmaCategory")
            count = self.import_amazon(base_dir, limit=options.get("limit"))
            self.stdout.write(self.style.SUCCESS(f"Imported {count} Amazon categories"))
        else:
            if options["truncate"]:
                WowCategory.objects.all().delete()
                self.stdout.write("Truncated WowCategory")
            count = self.import_wowma(base_dir, limit=options.get("limit"))
            self.stdout.write(self.style.SUCCESS(f"Imported {count} Wowma categories"))

    def import_amazon(self, base_dir: Path, limit=None) -> int:
        files = [
            f
            for f in sorted(base_dir.glob("*.xls*"))
            if f.suffix.lower() in (".xls", ".xlsx") and "Zone.Identifier" not in f.name
        ]
        if not files:
            raise CommandError(f"No BTG .xls/.xlsx files found in {base_dir}")

        objs = []
        total = 0
        for file in files:
            try:
                book = xlrd.open_workbook(file)
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"Skip {file.name}: {e}"))
                continue
            if book.nsheets < 2:
                continue
            sh = book.sheet_by_index(1)
            # Find header row containing "Node ID"
            header_row = None
            for r in range(min(10, sh.nrows)):
                row = [str(v).strip() for v in sh.row_values(r)]
                if any("node id" in v.lower() for v in row):
                    header_row = r
                    break
            if header_row is None:
                continue
            # Identify columns
            headers = [str(v).strip() for v in sh.row_values(header_row)]
            try:
                id_col = next(i for i, h in enumerate(headers) if h.lower().startswith("node id"))
                path_col = next(i for i, h in enumerate(headers) if "node path" in h.lower())
            except StopIteration:
                continue

            for r in range(header_row + 1, sh.nrows):
                if limit and total >= limit:
                    break
                row = sh.row_values(r)
                try:
                    node_id = int(str(row[id_col]).split(".")[0])
                except Exception:
                    continue
                node_path = str(row[path_col]).strip()
                if not node_path:
                    continue
                levels = [p.strip() for p in node_path.split("/") if p.strip()]
                fields = {
                    "product_cat_name": node_path,
                    "parent_cat_id": 0,
                    "qoo_cat_id": 0,
                    "wow_cat_id": 0,
                    "yahoo_cat_id": 0,
                }
                for idx, name in enumerate(levels[:8], start=1):
                    fields[f"level_{idx}_cat_name"] = name
                objs.append(AmaCategory(product_cat_id=node_id, **fields))
                total += 1
            if limit and total >= limit:
                break
        if objs:
            AmaCategory.objects.bulk_create(objs, batch_size=1000, ignore_conflicts=True)
        return total

    def import_wowma(self, base_dir: Path, limit=None) -> int:
        file = base_dir / "category.xlsx"
        if not file.exists():
            raise CommandError(f"Wowma category.xlsx not found in {base_dir}")
        wb = openpyxl.load_workbook(file, read_only=True)
        sh = wb.active
        # Map headers
        header = [str(c.value).strip() if c.value is not None else "" for c in next(sh.iter_rows(min_row=1, max_row=1))[0:]]
        try:
            id_idx = header.index("出品カテゴリID")
            full_idx = header.index("カテゴリフルパス")
            lvl1_idx = header.index("第一階層名")
            lvl2_idx = header.index("第二階層名")
            lvl3_idx = header.index("第三階層名")
            lvl4_idx = header.index("第四階層名")
        except ValueError:
            raise CommandError("Header row not as expected in Wowma category.xlsx")

        objs = []
        total = 0
        for row in sh.iter_rows(min_row=2):
            if limit and total >= limit:
                break
            values = [cell.value for cell in row]
            try:
                cat_id = int(str(values[id_idx]).split(".")[0])
            except Exception:
                continue
            full_path = str(values[full_idx]).strip() if values[full_idx] else ""
            lvl1 = str(values[lvl1_idx]).strip() if values[lvl1_idx] else ""
            lvl2 = str(values[lvl2_idx]).strip() if values[lvl2_idx] else ""
            lvl3 = str(values[lvl3_idx]).strip() if values[lvl3_idx] else ""
            lvl4 = str(values[lvl4_idx]).strip() if values[lvl4_idx] else ""
            objs.append(
                WowCategory(
                    product_cat_id=cat_id,
                    product_cat_name=full_path,
                    level_1_cat_name=lvl1,
                    level_2_cat_name=lvl2,
                    level_3_cat_name=lvl3,
                    level_4_cat_name=lvl4,
                )
            )
            total += 1
        if objs:
            WowCategory.objects.bulk_create(objs, batch_size=1000, ignore_conflicts=True)
        return total
